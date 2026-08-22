"""UC02 Excel administration for DI-owned configuration.

Owner-approved Excel domains:
- Document Types
- Extraction Profiles
- Requirement Profiles

The native DI tables remain authoritative. Upload only creates staging rows;
explicit confirmation creates/updates DRAFT/configurable native state. No WEF is
invented because these DI domains do not currently have an approved WEF concept.
"""
from __future__ import annotations

import hashlib
import io
import json
import uuid
from collections import defaultdict
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from verigence.di.document_ai.schemas import SCHEMA_REGISTRY
from verigence.di.repositories.tenants import provision_actor

DOCUMENT_TYPES = "DOCUMENT_TYPES"
EXTRACTION_PROFILES = "EXTRACTION_PROFILES"
REQUIREMENT_PROFILES = "REQUIREMENT_PROFILES"
EXCEL_MASTER_KEYS = frozenset({DOCUMENT_TYPES, EXTRACTION_PROFILES, REQUIREMENT_PROFILES})
TEMPLATE_VERSION = "1.0"

TEMPLATE_HEADERS: dict[str, tuple[str, ...]] = {
    DOCUMENT_TYPES: (
        "documentTypeKey",
        "displayName",
        "description",
        "physicalFormType",
        "requiresProcessing",
        "displayOrder",
    ),
    EXTRACTION_PROFILES: (
        "documentTypeKey",
        "profileName",
        "fieldKey",
        "enabled",
        "expected",
        "extractionInstruction",
        "scoreIncluded",
        "scoreWeight",
        "useForSubjectMatching",
        "subjectIdentifierType",
        "manualCorrectionAllowed",
        "displaySequence",
        "normalizerRuleKeys",
        "validatorRuleKeys",
        "validationSeverity",
    ),
    REQUIREMENT_PROFILES: (
        "profileKey",
        "description",
        "documentTypeKey",
        "requirementClassification",
        "minimumCount",
        "displaySequence",
    ),
}

_REQUIRED_HEADERS: dict[str, frozenset[str]] = {
    DOCUMENT_TYPES: frozenset({"documentTypeKey", "displayName", "physicalFormType"}),
    EXTRACTION_PROFILES: frozenset({"documentTypeKey", "profileName", "fieldKey"}),
    REQUIREMENT_PROFILES: frozenset({"profileKey", "documentTypeKey"}),
}


class ConfigImportError(ValueError):
    pass


class ConfigImportConflict(ConfigImportError):
    pass


def normalize_master_key(master_key: str) -> str:
    key = master_key.strip().upper().replace("-", "_")
    aliases = {
        "DOCUMENT_TYPE": DOCUMENT_TYPES,
        "DOCUMENT_TYPES": DOCUMENT_TYPES,
        "EXTRACTION_PROFILE": EXTRACTION_PROFILES,
        "EXTRACTION_PROFILES": EXTRACTION_PROFILES,
        "REQUIREMENT_PROFILE": REQUIREMENT_PROFILES,
        "REQUIREMENT_PROFILES": REQUIREMENT_PROFILES,
    }
    try:
        return aliases[key]
    except KeyError as exc:
        raise ConfigImportError("Unsupported DI Excel master") from exc


def build_template(master_key: str) -> bytes:
    key = normalize_master_key(master_key)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = key[:31]
    sheet.append(list(TEMPLATE_HEADERS[key]))
    if key == DOCUMENT_TYPES:
        sheet.append(["bank_statement", "Bank Statement", "", "PRINTABLE", True, 100])
    elif key == EXTRACTION_PROFILES:
        sheet.append([
            "bank_statement", "Bank Statement Extraction", "account_number",
            True, True, "Extract the account number", True, 1.0,
            True, "ACCOUNT_NUMBER", True, 10, "", "", "ERROR",
        ])
    else:
        sheet.append(["vehicle_sale", "Vehicle sale requirements", "bank_statement", "MANDATORY", 1, 10])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


async def stage_config_import(
    session: AsyncSession,
    *,
    tenant_id: str,
    master_key: str,
    idempotency_key: str,
    file_name: str,
    content: bytes,
    created_by_user_id: str,
) -> dict[str, Any]:
    key = normalize_master_key(master_key)
    if not file_name.lower().endswith(".xlsx"):
        raise ConfigImportError("Only .xlsx workbooks are accepted")
    if not content:
        raise ConfigImportError("Workbook is empty")
    file_hash = hashlib.sha256(content).hexdigest()

    existing = (
        await session.execute(
            text("""
                SELECT import_id, tenant_id, master_key, idempotency_key,
                       file_name, file_hash_sha256, template_version, status,
                       rows_parsed, valid_rows, warning_rows, error_rows,
                       result_reference, created_by_user_id, created_at_utc,
                       confirmed_by_user_id, confirmed_at_utc
                FROM docintel.config_imports
                WHERE tenant_id=:tenant_id AND idempotency_key=:idempotency_key
            """),
            {"tenant_id": tenant_id, "idempotency_key": idempotency_key},
        )
    ).mappings().one_or_none()
    if existing is not None:
        if existing["master_key"] != key or existing["file_hash_sha256"] != file_hash:
            raise ConfigImportConflict("Idempotency-Key was already used for a different import")
        return dict(existing)

    rows = _parse_workbook(content, key)
    validated: list[tuple[int, dict[str, Any], str, list[str]]] = []
    for row_number, parsed in rows:
        status, messages = await _validate_row(
            session,
            tenant_id=tenant_id,
            master_key=key,
            parsed=parsed,
        )
        validated.append((row_number, parsed, status, messages))
    _apply_cross_row_validation(key, validated)

    rows_parsed = len(validated)
    valid_rows = sum(1 for _, _, state, _ in validated if state == "VALID")
    warning_rows = sum(1 for _, _, state, _ in validated if state == "WARNING")
    error_rows = sum(1 for _, _, state, _ in validated if state == "ERROR")
    status = "VALIDATION_FAILED" if error_rows else "PREVIEW_READY"
    import_id = uuid.uuid4()
    now = datetime.now(UTC)
    await session.execute(
        text("""
            INSERT INTO docintel.config_imports (
                import_id, tenant_id, master_key, idempotency_key,
                file_name, file_hash_sha256, template_version, status,
                rows_parsed, valid_rows, warning_rows, error_rows,
                created_by_user_id, created_at_utc
            ) VALUES (
                :import_id, :tenant_id, :master_key, :idempotency_key,
                :file_name, :file_hash, :template_version, :status,
                :rows_parsed, :valid_rows, :warning_rows, :error_rows,
                :created_by, :now
            )
        """),
        {
            "import_id": import_id,
            "tenant_id": tenant_id,
            "master_key": key,
            "idempotency_key": idempotency_key,
            "file_name": file_name,
            "file_hash": file_hash,
            "template_version": TEMPLATE_VERSION,
            "status": status,
            "rows_parsed": rows_parsed,
            "valid_rows": valid_rows,
            "warning_rows": warning_rows,
            "error_rows": error_rows,
            "created_by": created_by_user_id,
            "now": now,
        },
    )
    for row_number, parsed, validation_status, messages in validated:
        await session.execute(
            text("""
                INSERT INTO docintel.config_import_rows (
                    tenant_id, import_id, row_number, parsed_data,
                    validation_status, validation_messages
                ) VALUES (
                    :tenant_id, :import_id, :row_number,
                    CAST(:parsed_data AS jsonb), :validation_status,
                    CAST(:messages AS jsonb)
                )
            """),
            {
                "tenant_id": tenant_id,
                "import_id": import_id,
                "row_number": row_number,
                "parsed_data": json.dumps(parsed),
                "validation_status": validation_status,
                "messages": json.dumps(messages),
            },
        )
    await session.commit()
    return await get_config_import(session, tenant_id=tenant_id, import_id=import_id)


async def get_config_import(
    session: AsyncSession,
    *,
    tenant_id: str,
    import_id: uuid.UUID,
) -> dict[str, Any]:
    row = (
        await session.execute(
            text("""
                SELECT import_id, tenant_id, master_key, idempotency_key,
                       file_name, file_hash_sha256, template_version, status,
                       rows_parsed, valid_rows, warning_rows, error_rows,
                       result_reference, created_by_user_id, created_at_utc,
                       confirmed_by_user_id, confirmed_at_utc
                FROM docintel.config_imports
                WHERE tenant_id=:tenant_id AND import_id=:import_id
            """),
            {"tenant_id": tenant_id, "import_id": import_id},
        )
    ).mappings().one_or_none()
    if row is None:
        raise ConfigImportError("Configuration import not found")
    return dict(row)


async def list_config_import_rows(
    session: AsyncSession,
    *,
    tenant_id: str,
    import_id: uuid.UUID,
) -> list[dict[str, Any]]:
    rows = (
        await session.execute(
            text("""
                SELECT row_number, parsed_data, validation_status, validation_messages
                FROM docintel.config_import_rows
                WHERE tenant_id=:tenant_id AND import_id=:import_id
                ORDER BY row_number
            """),
            {"tenant_id": tenant_id, "import_id": import_id},
        )
    ).mappings().all()
    return [dict(row) for row in rows]


async def cancel_config_import(
    session: AsyncSession,
    *,
    tenant_id: str,
    import_id: uuid.UUID,
) -> None:
    deleted = (
        await session.execute(
            text("""
                DELETE FROM docintel.config_imports
                WHERE tenant_id=:tenant_id AND import_id=:import_id
                  AND status IN ('PREVIEW_READY','VALIDATION_FAILED','FAILED')
                RETURNING import_id
            """),
            {"tenant_id": tenant_id, "import_id": import_id},
        )
    ).scalar_one_or_none()
    if deleted is None:
        raise ConfigImportConflict("Only unconfirmed configuration imports may be removed")
    await session.commit()


async def confirm_config_import(
    session: AsyncSession,
    *,
    tenant_id: str,
    import_id: uuid.UUID,
    confirmed_by_user_id: str,
) -> dict[str, Any]:
    header = await get_config_import(session, tenant_id=tenant_id, import_id=import_id)
    if header["status"] == "CONFIRMED":
        return header
    if header["status"] != "PREVIEW_READY" or header["error_rows"]:
        raise ConfigImportConflict("Import must be preview-ready with no blocking row errors")
    rows = await list_config_import_rows(session, tenant_id=tenant_id, import_id=import_id)
    await provision_actor(session, tenant_id, confirmed_by_user_id)
    key = str(header["master_key"])
    if key == DOCUMENT_TYPES:
        result_reference = await _confirm_document_types(
            session, tenant_id=tenant_id, rows=rows
        )
    elif key == EXTRACTION_PROFILES:
        result_reference = await _confirm_extraction_profiles(
            session,
            tenant_id=tenant_id,
            rows=rows,
            actor_id=confirmed_by_user_id,
        )
    elif key == REQUIREMENT_PROFILES:
        result_reference = await _confirm_requirement_profiles(
            session,
            tenant_id=tenant_id,
            rows=rows,
            actor_id=confirmed_by_user_id,
        )
    else:
        raise ConfigImportError("Unsupported DI Excel master")

    now = datetime.now(UTC)
    await session.execute(
        text("""
            UPDATE docintel.config_imports
            SET status='CONFIRMED', result_reference=CAST(:result AS jsonb),
                confirmed_by_user_id=:actor_id, confirmed_at_utc=:now
            WHERE tenant_id=:tenant_id AND import_id=:import_id
        """),
        {
            "tenant_id": tenant_id,
            "import_id": import_id,
            "result": json.dumps(result_reference),
            "actor_id": confirmed_by_user_id,
            "now": now,
        },
    )
    await session.commit()
    return await get_config_import(session, tenant_id=tenant_id, import_id=import_id)


def error_report_csv(rows: list[dict[str, Any]]) -> bytes:
    output = io.StringIO()
    output.write("rowNumber,validationStatus,messages\n")
    for row in rows:
        messages = " | ".join(str(item) for item in row.get("validation_messages") or [])
        escaped = messages.replace('"', '""')
        output.write(f'{row["row_number"]},{row["validation_status"]},"{escaped}"\n')
    return output.getvalue().encode("utf-8")


def _parse_workbook(content: bytes, master_key: str) -> list[tuple[int, dict[str, Any]]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ConfigImportError("Workbook could not be parsed") from exc
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        first = next(iterator)
    except StopIteration as exc:
        raise ConfigImportError("Workbook has no header row") from exc
    headers = [str(value).strip() if value is not None else "" for value in first]
    if len(headers) != len(set(headers)):
        raise ConfigImportError("Workbook contains duplicate header names")
    missing = _REQUIRED_HEADERS[master_key] - set(headers)
    if missing:
        raise ConfigImportError(f"Workbook is missing required columns: {', '.join(sorted(missing))}")
    allowed = set(TEMPLATE_HEADERS[master_key])
    unknown = {header for header in headers if header and header not in allowed}
    if unknown:
        raise ConfigImportError(f"Workbook contains unsupported columns: {', '.join(sorted(unknown))}")

    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(iterator, start=2):
        parsed = {
            header: _excel_scalar(value)
            for header, value in zip(headers, values, strict=False)
            if header
        }
        if not any(value not in (None, "") for value in parsed.values()):
            continue
        rows.append((row_number, parsed))
    if not rows:
        raise ConfigImportError("Workbook contains no data rows")
    return rows


def _excel_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (str, bool, int, float)):
        return value.strip() if isinstance(value, str) else value
    return str(value)


async def _validate_row(
    session: AsyncSession,
    *,
    tenant_id: str,
    master_key: str,
    parsed: dict[str, Any],
) -> tuple[str, list[str]]:
    errors: list[str] = []
    warnings: list[str] = []
    if master_key == DOCUMENT_TYPES:
        key = _required_text(parsed, "documentTypeKey", errors)
        _required_text(parsed, "displayName", errors)
        form_type = _required_text(parsed, "physicalFormType", errors).upper()
        if form_type and form_type not in {"GOVT_ID", "PRINTABLE", "HANDWRITTEN", "ADDITIONAL"}:
            errors.append("physicalFormType must be GOVT_ID, PRINTABLE, HANDWRITTEN or ADDITIONAL")
        _parse_bool(parsed.get("requiresProcessing"), "requiresProcessing", errors, default=True)
        _parse_positive_int(parsed.get("displayOrder"), "displayOrder", errors, default=100, allow_zero=True)
        if key and key.lower() != key:
            warnings.append("documentTypeKey should use the existing lowercase key convention")
    elif master_key == EXTRACTION_PROFILES:
        document_type_key = _required_text(parsed, "documentTypeKey", errors)
        _required_text(parsed, "profileName", errors)
        field_key = _required_text(parsed, "fieldKey", errors)
        await _validate_document_type_exists(session, tenant_id, document_type_key, errors)
        await _validate_canonical_field_exists(session, tenant_id, field_key, errors)
        _parse_bool(parsed.get("enabled"), "enabled", errors, default=True)
        _parse_bool(parsed.get("expected"), "expected", errors, default=True)
        _parse_bool(parsed.get("scoreIncluded"), "scoreIncluded", errors, default=True)
        _parse_bool(parsed.get("useForSubjectMatching"), "useForSubjectMatching", errors, default=False)
        _parse_bool(parsed.get("manualCorrectionAllowed"), "manualCorrectionAllowed", errors, default=True)
        _parse_nonnegative_float(parsed.get("scoreWeight"), "scoreWeight", errors, default=1.0)
        _parse_positive_int(parsed.get("displaySequence"), "displaySequence", errors, default=100, allow_zero=True)
        await _validate_rule_keys(session, "normalization_rule_catalog", parsed.get("normalizerRuleKeys"), errors)
        await _validate_rule_keys(session, "validation_rule_catalog", parsed.get("validatorRuleKeys"), errors)
        severity = str(parsed.get("validationSeverity") or "ERROR").upper()
        if severity not in {"INFO", "WARNING", "ERROR"}:
            errors.append("validationSeverity must be INFO, WARNING or ERROR")
        schema = SCHEMA_REGISTRY.get(document_type_key)
        if schema is not None and field_key:
            registry_fields = {field.key for field in schema.fields}
            if field_key not in registry_fields:
                warnings.append(
                    "fieldKey is not present in the D25 Python schema registry for this Document Type"
                )
    else:
        _required_text(parsed, "profileKey", errors)
        document_type_key = _required_text(parsed, "documentTypeKey", errors)
        await _validate_document_type_exists(session, tenant_id, document_type_key, errors)
        classification = str(parsed.get("requirementClassification") or "MANDATORY").upper()
        if classification not in {"MANDATORY", "OPTIONAL"}:
            errors.append("requirementClassification must be MANDATORY or OPTIONAL")
        _parse_positive_int(parsed.get("minimumCount"), "minimumCount", errors, default=1)
        _parse_positive_int(parsed.get("displaySequence"), "displaySequence", errors, default=100, allow_zero=True)
    if errors:
        return "ERROR", errors + warnings
    if warnings:
        return "WARNING", warnings
    return "VALID", []


def _apply_cross_row_validation(
    master_key: str,
    rows: list[tuple[int, dict[str, Any], str, list[str]]],
) -> None:
    seen: dict[tuple[str, ...], int] = {}
    for index, (_, parsed, _state, messages) in enumerate(rows):
        identity: tuple[str, ...]
        if master_key == DOCUMENT_TYPES:
            identity = (str(parsed.get("documentTypeKey") or ""),)
        elif master_key == EXTRACTION_PROFILES:
            identity = (
                str(parsed.get("documentTypeKey") or ""),
                str(parsed.get("profileName") or ""),
                str(parsed.get("fieldKey") or ""),
            )
        else:
            identity = (
                str(parsed.get("profileKey") or ""),
                str(parsed.get("documentTypeKey") or ""),
            )
        if identity in seen:
            messages.append(f"Duplicate logical row; first occurrence is workbook row {seen[identity]}")
            rows[index] = (rows[index][0], parsed, "ERROR", messages)
        else:
            seen[identity] = rows[index][0]

    if master_key == EXTRACTION_PROFILES:
        groups: dict[tuple[str, str], list[int]] = defaultdict(list)
        for index, (_, parsed, _, _) in enumerate(rows):
            groups[(str(parsed.get("documentTypeKey") or ""), str(parsed.get("profileName") or ""))].append(index)
        for indexes in groups.values():
            any_scored = False
            for index in indexes:
                parsed = rows[index][1]
                errors: list[str] = []
                included = _parse_bool(parsed.get("scoreIncluded"), "scoreIncluded", errors, default=True)
                weight = _parse_nonnegative_float(parsed.get("scoreWeight"), "scoreWeight", errors, default=1.0)
                enabled = _parse_bool(parsed.get("enabled"), "enabled", errors, default=True)
                expected = _parse_bool(parsed.get("expected"), "expected", errors, default=True)
                any_scored = any_scored or bool(enabled and expected and included and weight > 0)
            if not any_scored:
                for index in indexes:
                    row_no, parsed, _, messages = rows[index]
                    rows[index] = (
                        row_no,
                        parsed,
                        "ERROR",
                        messages + ["Extraction Profile requires at least one enabled expected scored field with weight > 0"],
                    )


def _required_text(parsed: dict[str, Any], key: str, errors: list[str]) -> str:
    value = parsed.get(key)
    if not isinstance(value, str) or not value.strip():
        errors.append(f"{key} is required")
        return ""
    return value.strip()


def _parse_bool(value: Any, key: str, errors: list[str], *, default: bool) -> bool:
    if value in (None, ""):
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and value in {0, 1}:
        return bool(value)
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"true", "yes", "y", "1"}:
            return True
        if normalized in {"false", "no", "n", "0"}:
            return False
    errors.append(f"{key} must be a boolean")
    return default


def _parse_positive_int(
    value: Any,
    key: str,
    errors: list[str],
    *,
    default: int,
    allow_zero: bool = False,
) -> int:
    if value in (None, ""):
        return default
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        errors.append(f"{key} must be an integer")
        return default
    if parsed < 0 or (parsed == 0 and not allow_zero):
        errors.append(f"{key} must be {'non-negative' if allow_zero else 'positive'}")
    return parsed


def _parse_nonnegative_float(value: Any, key: str, errors: list[str], *, default: float) -> float:
    if value in (None, ""):
        return default
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        errors.append(f"{key} must be numeric")
        return default
    if parsed < 0:
        errors.append(f"{key} must be non-negative")
    return parsed


async def _validate_document_type_exists(
    session: AsyncSession, tenant_id: str, key: str, errors: list[str]
) -> None:
    if not key:
        return
    found = (
        await session.execute(
            text("""
                SELECT 1 FROM docintel.document_types
                WHERE document_type_key=:key
                  AND status != 'RETIRED'
                  AND (owner_tenant_id=:tenant_id OR owner_tenant_id IS NULL)
                LIMIT 1
            """),
            {"key": key, "tenant_id": tenant_id},
        )
    ).scalar_one_or_none()
    if found is None:
        errors.append(f"Unknown Document Type key: {key}")


async def _validate_canonical_field_exists(
    session: AsyncSession, tenant_id: str, key: str, errors: list[str]
) -> None:
    if not key:
        return
    found = (
        await session.execute(
            text("""
                SELECT 1 FROM docintel.canonical_fields
                WHERE field_key=:key AND status='ACTIVE'
                  AND (owner_tenant_id=:tenant_id OR owner_tenant_id IS NULL)
                LIMIT 1
            """),
            {"key": key, "tenant_id": tenant_id},
        )
    ).scalar_one_or_none()
    if found is None:
        errors.append(f"Unknown canonical field key: {key}")


async def _validate_rule_keys(
    session: AsyncSession,
    table: str,
    value: Any,
    errors: list[str],
) -> None:
    if table not in {"normalization_rule_catalog", "validation_rule_catalog"}:
        raise ValueError("Unsupported rule catalogue")
    for key in _split_rule_keys(value):
        found = (
            await session.execute(
                text(f"SELECT 1 FROM docintel.{table} WHERE rule_key=:key AND status='ACTIVE'"),
                {"key": key},
            )
        ).scalar_one_or_none()
        if found is None:
            errors.append(f"Unknown active rule key: {key}")


def _split_rule_keys(value: Any) -> list[str]:
    if value in (None, ""):
        return []
    return [part.strip() for part in str(value).replace(";", ",").split(",") if part.strip()]


async def _effective_document_type_id(session: AsyncSession, tenant_id: str, key: str) -> uuid.UUID:
    value = (
        await session.execute(
            text("""
                SELECT document_type_id
                FROM docintel.document_types
                WHERE document_type_key=:key
                  AND status != 'RETIRED'
                  AND (owner_tenant_id=:tenant_id OR owner_tenant_id IS NULL)
                ORDER BY CASE WHEN owner_tenant_id=:tenant_id THEN 0 ELSE 1 END
                LIMIT 1
            """),
            {"key": key, "tenant_id": tenant_id},
        )
    ).scalar_one()
    return uuid.UUID(str(value))


async def _canonical_field_id(session: AsyncSession, tenant_id: str, key: str) -> uuid.UUID:
    value = (
        await session.execute(
            text("""
                SELECT canonical_field_id
                FROM docintel.canonical_fields
                WHERE field_key=:key AND status='ACTIVE'
                  AND (owner_tenant_id=:tenant_id OR owner_tenant_id IS NULL)
                ORDER BY CASE WHEN owner_tenant_id=:tenant_id THEN 0 ELSE 1 END
                LIMIT 1
            """),
            {"key": key, "tenant_id": tenant_id},
        )
    ).scalar_one()
    return uuid.UUID(str(value))


async def _confirm_document_types(
    session: AsyncSession,
    *,
    tenant_id: str,
    rows: list[dict[str, Any]],
) -> dict[str, Any]:
    configured: list[str] = []
    now = datetime.now(UTC)
    for row in rows:
        data = row["parsed_data"]
        key = str(data["documentTypeKey"]).strip()
        tenant_dt = (
            await session.execute(
                text("""
                    SELECT document_type_id, status
                    FROM docintel.document_types
                    WHERE owner_tenant_id=:tenant_id AND document_type_key=:key
                """),
                {"tenant_id": tenant_id, "key": key},
            )
        ).one_or_none()
        if tenant_dt is None:
            global_dt = (
                await session.execute(
                    text("""
                        SELECT document_type_id FROM docintel.document_types
                        WHERE owner_tenant_id IS NULL AND document_type_key=:key
                          AND status != 'RETIRED'
                    """),
                    {"key": key},
                )
            ).scalar_one_or_none()
            if global_dt is None:
                document_type_id = uuid.uuid4()
                await session.execute(
                    text("""
                        INSERT INTO docintel.document_types (
                            document_type_id, owner_tenant_id, document_type_key,
                            display_name, description, category, status,
                            created_at_utc, updated_at_utc
                        ) VALUES (
                            :id, :tenant_id, :key, :display_name, :description,
                            :category, 'DRAFT', :now, :now
                        )
                    """),
                    {
                        "id": document_type_id,
                        "tenant_id": tenant_id,
                        "key": key,
                        "display_name": str(data["displayName"]).strip(),
                        "description": data.get("description"),
                        "category": str(data.get("physicalFormType") or "ADDITIONAL").upper(),
                        "now": now,
                    },
                )
            else:
                document_type_id = global_dt
        else:
            document_type_id = tenant_dt[0]
            await session.execute(
                text("""
                    UPDATE docintel.document_types
                    SET display_name=:display_name, description=:description,
                        category=:category, updated_at_utc=:now
                    WHERE document_type_id=:id AND status != 'RETIRED'
                """),
                {
                    "id": document_type_id,
                    "display_name": str(data["displayName"]).strip(),
                    "description": data.get("description"),
                    "category": str(data.get("physicalFormType") or "ADDITIONAL").upper(),
                    "now": now,
                },
            )
        bool_errors: list[str] = []
        requires_processing = _parse_bool(
            data.get("requiresProcessing"), "requiresProcessing", bool_errors, default=True
        )
        order_errors: list[str] = []
        display_order = _parse_positive_int(
            data.get("displayOrder"), "displayOrder", order_errors, default=100, allow_zero=True
        )
        await session.execute(
            text("""
                INSERT INTO docintel.tenant_document_types (
                    tenant_id, document_type_id, physical_form_type,
                    requires_processing, is_active, display_order,
                    created_at_utc, updated_at_utc
                ) VALUES (
                    :tenant_id, :document_type_id, :form_type,
                    :requires_processing, true, :display_order, :now, :now
                )
                ON CONFLICT (tenant_id, document_type_id) DO UPDATE
                SET physical_form_type=EXCLUDED.physical_form_type,
                    requires_processing=EXCLUDED.requires_processing,
                    is_active=true,
                    display_order=EXCLUDED.display_order,
                    updated_at_utc=EXCLUDED.updated_at_utc
            """),
            {
                "tenant_id": tenant_id,
                "document_type_id": document_type_id,
                "form_type": str(data.get("physicalFormType") or "ADDITIONAL").upper(),
                "requires_processing": requires_processing,
                "display_order": display_order,
                "now": now,
            },
        )
        configured.append(key)
    return {"documentTypeKeys": configured}


async def _confirm_extraction_profiles(
    session: AsyncSession,
    *,
    tenant_id: str,
    rows: list[dict[str, Any]],
    actor_id: str,
) -> dict[str, Any]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        data = row["parsed_data"]
        grouped[(str(data["documentTypeKey"]), str(data["profileName"]))].append(data)
    profile_ids: list[str] = []
    now = datetime.now(UTC)
    for (document_type_key, profile_name), items in grouped.items():
        dt_id = await _effective_document_type_id(session, tenant_id, document_type_key)
        version_no = (
            await session.execute(
                text("""
                    SELECT COALESCE(MAX(version_no), 0) + 1
                    FROM docintel.extraction_profiles
                    WHERE document_type_id=:dt_id AND scope_tenant_id=:tenant_id
                """),
                {"dt_id": dt_id, "tenant_id": tenant_id},
            )
        ).scalar_one()
        profile_id = uuid.uuid4()
        await session.execute(
            text("""
                INSERT INTO docintel.extraction_profiles (
                    profile_id, document_type_id, scope_tenant_id, version_no,
                    profile_name, status, created_by_actor_id,
                    created_at_utc, updated_at_utc
                ) VALUES (
                    :profile_id, :dt_id, :tenant_id, :version_no,
                    :profile_name, 'DRAFT', :actor_id, :now, :now
                )
            """),
            {
                "profile_id": profile_id,
                "dt_id": dt_id,
                "tenant_id": tenant_id,
                "version_no": version_no,
                "profile_name": profile_name,
                "actor_id": actor_id,
                "now": now,
            },
        )
        for data in items:
            field_id = await _canonical_field_id(session, tenant_id, str(data["fieldKey"]))
            profile_field_id = uuid.uuid4()
            errors: list[str] = []
            await session.execute(
                text("""
                    INSERT INTO docintel.extraction_profile_fields (
                        profile_field_id, profile_id, canonical_field_id,
                        enabled, expected, extraction_instruction,
                        score_included, score_weight, use_for_subject_matching,
                        subject_identifier_type, manual_correction_allowed,
                        display_sequence, created_at_utc, updated_at_utc
                    ) VALUES (
                        :field_id, :profile_id, :canonical_field_id,
                        :enabled, :expected, :instruction,
                        :score_included, :score_weight, :subject_matching,
                        :identifier_type, :manual_allowed,
                        :display_sequence, :now, :now
                    )
                """),
                {
                    "field_id": profile_field_id,
                    "profile_id": profile_id,
                    "canonical_field_id": field_id,
                    "enabled": _parse_bool(data.get("enabled"), "enabled", errors, default=True),
                    "expected": _parse_bool(data.get("expected"), "expected", errors, default=True),
                    "instruction": data.get("extractionInstruction"),
                    "score_included": _parse_bool(data.get("scoreIncluded"), "scoreIncluded", errors, default=True),
                    "score_weight": _parse_nonnegative_float(data.get("scoreWeight"), "scoreWeight", errors, default=1.0),
                    "subject_matching": _parse_bool(data.get("useForSubjectMatching"), "useForSubjectMatching", errors, default=False),
                    "identifier_type": data.get("subjectIdentifierType") or None,
                    "manual_allowed": _parse_bool(data.get("manualCorrectionAllowed"), "manualCorrectionAllowed", errors, default=True),
                    "display_sequence": _parse_positive_int(data.get("displaySequence"), "displaySequence", errors, default=100, allow_zero=True),
                    "now": now,
                },
            )
            for sequence_no, rule_key in enumerate(_split_rule_keys(data.get("normalizerRuleKeys")), start=1):
                await session.execute(
                    text("""
                        INSERT INTO docintel.profile_field_normalizers (
                            profile_field_normalizer_id, profile_field_id,
                            sequence_no, rule_key
                        ) VALUES (:id, :field_id, :sequence_no, :rule_key)
                    """),
                    {
                        "id": uuid.uuid4(),
                        "field_id": profile_field_id,
                        "sequence_no": sequence_no,
                        "rule_key": rule_key,
                    },
                )
            severity = str(data.get("validationSeverity") or "ERROR").upper()
            for sequence_no, rule_key in enumerate(_split_rule_keys(data.get("validatorRuleKeys")), start=1):
                await session.execute(
                    text("""
                        INSERT INTO docintel.profile_field_validators (
                            profile_field_validator_id, profile_field_id,
                            sequence_no, rule_key, severity
                        ) VALUES (:id, :field_id, :sequence_no, :rule_key, :severity)
                    """),
                    {
                        "id": uuid.uuid4(),
                        "field_id": profile_field_id,
                        "sequence_no": sequence_no,
                        "rule_key": rule_key,
                        "severity": severity,
                    },
                )
        profile_ids.append(str(profile_id))
    return {"draftProfileIds": profile_ids}


async def _confirm_requirement_profiles(
    session: AsyncSession,
    *,
    tenant_id: str,
    rows: list[dict[str, Any]],
    actor_id: str,
) -> dict[str, Any]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        data = row["parsed_data"]
        grouped[str(data["profileKey"])].append(data)
    profile_ids: list[str] = []
    now = datetime.now(UTC)
    for profile_key, items in grouped.items():
        version_no = (
            await session.execute(
                text("""
                    SELECT COALESCE(MAX(version_no), 0) + 1
                    FROM docintel.document_requirement_profiles
                    WHERE tenant_id=:tenant_id AND profile_key=:profile_key
                """),
                {"tenant_id": tenant_id, "profile_key": profile_key},
            )
        ).scalar_one()
        profile_id = uuid.uuid4()
        await session.execute(
            text("""
                INSERT INTO docintel.document_requirement_profiles (
                    tenant_id, requirement_profile_id, profile_key, version_no,
                    description, status, created_by_actor_id,
                    created_at_utc, updated_at_utc
                ) VALUES (
                    :tenant_id, :profile_id, :profile_key, :version_no,
                    :description, 'DRAFT', :actor_id, :now, :now
                )
            """),
            {
                "tenant_id": tenant_id,
                "profile_id": profile_id,
                "profile_key": profile_key,
                "version_no": version_no,
                "description": items[0].get("description"),
                "actor_id": actor_id,
                "now": now,
            },
        )
        for data in items:
            errors: list[str] = []
            dt_id = await _effective_document_type_id(session, tenant_id, str(data["documentTypeKey"]))
            await session.execute(
                text("""
                    INSERT INTO docintel.document_requirement_profile_items (
                        tenant_id, requirement_item_id, requirement_profile_id,
                        document_type_id, requirement_classification,
                        minimum_count, display_sequence, enabled, created_at_utc
                    ) VALUES (
                        :tenant_id, :item_id, :profile_id, :document_type_id,
                        :classification, :minimum_count, :display_sequence,
                        true, :now
                    )
                """),
                {
                    "tenant_id": tenant_id,
                    "item_id": uuid.uuid4(),
                    "profile_id": profile_id,
                    "document_type_id": dt_id,
                    "classification": str(data.get("requirementClassification") or "MANDATORY").upper(),
                    "minimum_count": _parse_positive_int(data.get("minimumCount"), "minimumCount", errors, default=1),
                    "display_sequence": _parse_positive_int(data.get("displaySequence"), "displaySequence", errors, default=100, allow_zero=True),
                    "now": now,
                },
            )
        profile_ids.append(str(profile_id))
    return {"draftRequirementProfileIds": profile_ids}
