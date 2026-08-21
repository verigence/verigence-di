from __future__ import annotations

import io
import uuid

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import text

from verigence.di.api.v1.project_masters import (
    _MASTER_CATALOG,
    _list_versions,
    _publish_version,
)
from verigence.di.application.config_imports import (
    DOCUMENT_TYPES,
    EXTRACTION_PROFILES,
    REQUIREMENT_PROFILES,
    TEMPLATE_HEADERS,
    ConfigImportError,
    build_template,
    confirm_config_import,
    list_config_import_rows,
    normalize_master_key,
    stage_config_import,
)
from verigence.di.repositories.database import set_tenant_context
from verigence.di.repositories.tenants import provision_retention_policy, provision_tenant


def _workbook(master_key: str, row: list[object]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(list(TEMPLATE_HEADERS[master_key]))
    sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


@pytest.mark.no_docker
def test_master_catalog_keeps_di_excel_domains_without_wef() -> None:
    assert set(_MASTER_CATALOG) == {
        DOCUMENT_TYPES,
        EXTRACTION_PROFILES,
        REQUIREMENT_PROFILES,
    }
    for item in _MASTER_CATALOG.values():
        assert item["ownerModule"] == "DI"
        assert item["administrationModes"] == ["FORM", "EXCEL"]
        assert item["requiresWEF"] is False


@pytest.mark.no_docker
@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ("document-type", DOCUMENT_TYPES),
        ("DOCUMENT_TYPES", DOCUMENT_TYPES),
        ("extraction_profile", EXTRACTION_PROFILES),
        ("requirement-profiles", REQUIREMENT_PROFILES),
    ],
)
def test_normalize_master_key_aliases(raw: str, expected: str) -> None:
    assert normalize_master_key(raw) == expected


@pytest.mark.no_docker
def test_normalize_master_key_rejects_unknown_domain() -> None:
    with pytest.raises(ConfigImportError):
        normalize_master_key("quality-policy")


@pytest.mark.no_docker
@pytest.mark.parametrize(
    "master_key",
    [DOCUMENT_TYPES, EXTRACTION_PROFILES, REQUIREMENT_PROFILES],
)
def test_templates_use_the_approved_native_headers(master_key: str) -> None:
    workbook = load_workbook(io.BytesIO(build_template(master_key)), read_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows())]
    assert headers == list(TEMPLATE_HEADERS[master_key])
    assert "WEF" not in headers
    assert "Valid From" not in headers


@pytest.mark.asyncio
async def test_document_type_excel_confirm_and_publish(db_session) -> None:  # type: ignore[no-untyped-def]
    tenant_id = f"uc02-master-{uuid.uuid4().hex[:8]}"
    actor_id = str(uuid.uuid4())
    document_type_key = f"uc02_type_{uuid.uuid4().hex[:8]}"

    await set_tenant_context(db_session, tenant_id)
    await provision_tenant(db_session, tenant_id)
    await provision_retention_policy(db_session, tenant_id)
    await db_session.commit()

    content = _workbook(
        DOCUMENT_TYPES,
        [
            document_type_key,
            "UC02 Document",
            "Project onboarding test document",
            "PRINTABLE",
            True,
            10,
        ],
    )
    header = await stage_config_import(
        db_session,
        tenant_id=tenant_id,
        master_key=DOCUMENT_TYPES,
        idempotency_key=f"import-{uuid.uuid4()}",
        file_name="document-types.xlsx",
        content=content,
        created_by_user_id=actor_id,
    )
    assert header["status"] == "PREVIEW_READY"
    rows = await list_config_import_rows(
        db_session, tenant_id=tenant_id, import_id=header["import_id"]
    )
    assert rows[0]["validation_status"] in {"VALID", "WARNING"}

    confirmed = await confirm_config_import(
        db_session,
        tenant_id=tenant_id,
        import_id=header["import_id"],
        confirmed_by_user_id=actor_id,
    )
    assert confirmed["status"] == "CONFIRMED"

    version_id = (
        await db_session.execute(
            text(
                """
                SELECT document_type_id
                FROM docintel.document_types
                WHERE owner_tenant_id=:tenant_id AND document_type_key=:key
                """
            ),
            {"tenant_id": tenant_id, "key": document_type_key},
        )
    ).scalar_one()
    await _publish_version(
        db_session,
        tenant_id=tenant_id,
        master_key=DOCUMENT_TYPES,
        version_id=version_id,
        actor_id=actor_id,
    )
    await db_session.commit()

    versions = await _list_versions(
        db_session, tenant_id=tenant_id, master_key=DOCUMENT_TYPES
    )
    published = next(item for item in versions if item["versionId"] == version_id)
    assert published["status"] == "ACTIVE"
    assert published["activeForTenant"] is True


@pytest.mark.asyncio
async def test_extraction_and_requirement_excel_create_drafts_and_publish(db_session) -> None:  # type: ignore[no-untyped-def]
    tenant_id = f"uc02-profile-{uuid.uuid4().hex[:8]}"
    actor_id = str(uuid.uuid4())
    document_type_key = f"uc02_profile_type_{uuid.uuid4().hex[:8]}"
    field_key = f"uc02_field_{uuid.uuid4().hex[:8]}"

    await set_tenant_context(db_session, tenant_id)
    await provision_tenant(db_session, tenant_id)
    await provision_retention_policy(db_session, tenant_id)
    now = __import__("datetime").datetime.now(__import__("datetime").UTC)
    document_type_id = uuid.uuid4()
    canonical_field_id = uuid.uuid4()
    await db_session.execute(
        text(
            """
            INSERT INTO docintel.document_types (
                document_type_id, owner_tenant_id, document_type_key,
                display_name, category, status, created_at_utc, updated_at_utc
            ) VALUES (
                :id, :tenant_id, :key, 'UC02 Profile Type', 'PRINTABLE',
                'ACTIVE', :now, :now
            )
            """
        ),
        {
            "id": document_type_id,
            "tenant_id": tenant_id,
            "key": document_type_key,
            "now": now,
        },
    )
    await db_session.execute(
        text(
            """
            INSERT INTO docintel.canonical_fields (
                canonical_field_id, owner_tenant_id, field_key, display_name,
                data_type, status, created_at_utc, updated_at_utc
            ) VALUES (
                :id, :tenant_id, :key, 'UC02 Field', 'STRING',
                'ACTIVE', :now, :now
            )
            """
        ),
        {
            "id": canonical_field_id,
            "tenant_id": tenant_id,
            "key": field_key,
            "now": now,
        },
    )
    await db_session.commit()

    extraction_content = _workbook(
        EXTRACTION_PROFILES,
        [
            document_type_key,
            "UC02 Extraction",
            field_key,
            True,
            True,
            "Extract the UC02 field",
            True,
            1.0,
            False,
            "",
            True,
            10,
            "",
            "",
            "ERROR",
        ],
    )
    extraction_import = await stage_config_import(
        db_session,
        tenant_id=tenant_id,
        master_key=EXTRACTION_PROFILES,
        idempotency_key=f"extract-{uuid.uuid4()}",
        file_name="extraction.xlsx",
        content=extraction_content,
        created_by_user_id=actor_id,
    )
    assert extraction_import["error_rows"] == 0
    extraction_confirmed = await confirm_config_import(
        db_session,
        tenant_id=tenant_id,
        import_id=extraction_import["import_id"],
        confirmed_by_user_id=actor_id,
    )
    extraction_id = uuid.UUID(
        extraction_confirmed["result_reference"]["draftProfileIds"][0]
    )
    await _publish_version(
        db_session,
        tenant_id=tenant_id,
        master_key=EXTRACTION_PROFILES,
        version_id=extraction_id,
        actor_id=actor_id,
    )

    requirement_content = _workbook(
        REQUIREMENT_PROFILES,
        [
            "uc02-requirements",
            "UC02 requirements",
            document_type_key,
            "MANDATORY",
            1,
            10,
        ],
    )
    requirement_import = await stage_config_import(
        db_session,
        tenant_id=tenant_id,
        master_key=REQUIREMENT_PROFILES,
        idempotency_key=f"req-{uuid.uuid4()}",
        file_name="requirements.xlsx",
        content=requirement_content,
        created_by_user_id=actor_id,
    )
    assert requirement_import["error_rows"] == 0
    requirement_confirmed = await confirm_config_import(
        db_session,
        tenant_id=tenant_id,
        import_id=requirement_import["import_id"],
        confirmed_by_user_id=actor_id,
    )
    requirement_id = uuid.UUID(
        requirement_confirmed["result_reference"]["draftRequirementProfileIds"][0]
    )
    await _publish_version(
        db_session,
        tenant_id=tenant_id,
        master_key=REQUIREMENT_PROFILES,
        version_id=requirement_id,
        actor_id=actor_id,
    )
    await db_session.commit()

    extraction_versions = await _list_versions(
        db_session, tenant_id=tenant_id, master_key=EXTRACTION_PROFILES
    )
    requirement_versions = await _list_versions(
        db_session, tenant_id=tenant_id, master_key=REQUIREMENT_PROFILES
    )
    assert next(
        item for item in extraction_versions if item["versionId"] == extraction_id
    )["status"] == "PUBLISHED"
    assert next(
        item for item in requirement_versions if item["versionId"] == requirement_id
    )["status"] == "PUBLISHED"
